import openpyxl as op
from openpyxl.styles.numbers import BUILTIN_FORMATS

path = r'C:\generation_results\new.xlsx'
resurse_path = r'C:\generation_results\fin.xlsx'


def analise_column(file, column_inx, min_row):
    """
    Функция обрабатывает заданные колонки, находит "0", не целочисленных значений
    и текст, затем удаляет обнаруженные несоответвия и округляя числа типа float.

    :param file: маршрут к файлам, которые содеражт обрабатываемые колонки
    :param column_inx: передается список из идексов колонок
    :param min_row: задаем стартовую строку
    """

    workbook = op.load_workbook(file)
    sheet = workbook.active
    for col in column_inx:
        for row in sheet.iter_cols(min_col=col, max_col=col, min_row=min_row):
            for cell in row:
                # Отбираем все значения кроме None
                if type(cell.value) is str:
                    try:
                        # преобразуем в float, если не получилось, значит имеем дело не с числом
                        float(cell.value)
                    except ValueError:
                        # если не получислось преобразовать в float, то
                        # находим кординаты ячейки и номер строки,
                        coordinate_of_cells = cell.coordinate
                        number_of_rows = cell.row
                        sheet[coordinate_of_cells] = None
                        # очищаем некорректную ячейку и полностью удаляем строку
                        sheet.delete_rows(number_of_rows)
                    if type(cell.value) is str and '.' in cell.value:
                        # округляем очищенные данные
                        coordinate_of_float_cells = cell.coordinate
                        float_format = float(cell.value)
                        sheet[coordinate_of_float_cells] = round(float_format)
                # отбираем все значения равные нулю
                if cell.value == '0' or cell.value == 0:
                    # определить координаты этих ячеек
                    coordinate_of_cell = cell.coordinate
                    # передать эти координаты на обнуление ниже
                    sheet[coordinate_of_cell] = None
                    # записать обновленные данные в эксель
            workbook.save(path)
    return 'Проверка закончена'


# print(analise_column(resurse_path, [13, 14, 15], 2))


column_inx = [6, 23]
workbook = op.load_workbook(r'C:\generation_results\fin.xlsx')
sheet = workbook.active
for col in column_inx:
    for row in sheet.iter_cols(min_col=col, max_col=col, min_row=2):
        for cell in row:
            try:
                # преобразуем в datetime, если не получилось, значит имеем дело не с числом
                coordinate_of_cells = cell.coordinate
                sheet[coordinate_of_cells].number_format = BUILTIN_FORMATS[14]
            except ValueError:
                # если не получислось преобразовать в datetime, то
                # находим кординаты ячейки и номер строки и
                # очищаем некорректную ячейку
                coordinate_of_cells = cell.coordinate
                sheet[coordinate_of_cells] = None

        workbook.save(path)

