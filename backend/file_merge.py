import os
from datetime import date

import pandas as pd
import openpyxl as op
from pyexcelerate import Workbook

from backend.check_system import head_of_table
from backend.utils import timemometr
from settings import group_dir, xlsx_file, file_path


# @timemometr
def file_dir_merge(sheet_name: str = 'Лист1'):
    """Функция объединяет файлы ориентируясь на числовую шапку файла.

    :param sheet_name: имя лста
    :return: 'текст'
    """
    current_datetime = date.today()
    single_file_list = []
    for file in os.listdir(fr'C:\{group_dir}\{xlsx_file}'):
        wb = op.load_workbook(fr'C:\{group_dir}\{xlsx_file}\{file}',
                              read_only=True,
                              data_only=True)
        # проверяем лист на его корректность
        try:
            ws = wb[sheet_name]
        # если указанного листа не существует, переходим на активный лист
        except KeyError:
            ws = wb.active
        try:
            head = head_of_table(fr'C:\{group_dir}\{xlsx_file}\{file}', sheet_name)
        except Exception as error:
            return f'Ошибка: {error}. ' \
                   f'Не получается прочитать файл!'
        for row_col_ind, head in head.items():
            for ind, row in enumerate(ws.iter_rows(min_col=row_col_ind[1],
                                                   max_col=36,
                                                   min_row=row_col_ind[0] + 1,
                                                   values_only=True)):
                single_file_list.append(row)

    main_df = pd.DataFrame(single_file_list,
                           dtype='object')
    values = [main_df.columns] + list(main_df.values)
    wb = Workbook()
    wb.new_sheet('Финал', data=values)
    wb.save(fr'C:\generation_results\НБО_свод_{current_datetime.strftime("%d.%m.%Y")}.xlsx')
    return 'Слияние выполнено!'


# print(file_dir_merge())


# def new_reader_excel(path):
#     pass
