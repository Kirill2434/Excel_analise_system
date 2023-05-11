import os
import glob

import openpyxl as op
import pandas as pd
from pathlib import Path

from pandas import ExcelWriter

# from settings import file_path


def timemometr(func):
    from time import time

    def wrapper(*args):
        start_time = time()
        value = func(*args)
        end_time = time()
        print(f'Время выполнения функции {end_time-start_time} сек.')
        return value
    return wrapper


# @timemometr
def record_to_excel(obj, report_name, sheet_name):
    """Универсальная функция записи результатов выполнения различных
    проверок и действий в других модулях проекта.

    @param obj: датафрейм который необходимо записать в файл формата xlsx
    @param sheet_name: имя листа на который будут записаны данные
    @param report_name: имя отчетного файла
    """

    # если файл-отчет отсувует в папке, то записываем первый файл-отчет
    try:
        if os.path.exists(fr'C:\generation_results\{report_name}.xlsx') is False:
            try:
                with ExcelWriter(fr'C:\generation_results\{report_name}.xlsx',
                                 engine='openpyxl') as writer:
                    # проверяем тип объекта
                    # либо список
                    if isinstance(obj, list):
                        df = pd.DataFrame(obj).transpose()
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    # либо словарь
                    if isinstance(obj, dict):
                        df = pd.DataFrame.from_dict(obj, orient='index')
                        df = df.transpose()
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            except IndexError:
                obj.to_excel(fr'C:\generation_results\{report_name}.xlsx',
                             sheet_name=sheet_name,
                             index=False)
        # когда файл уже есть:
        else:
            # счиаем кол-во файлов-отчетов в папке
            sum_of_files = len(list(glob.glob(fr'C:\generation_results\{report_name}??.xlsx')))
            num = sum_of_files + 1
            # создаем новое имя файла с учетом нового номера
            new_file_name = fr'C:\generation_results\{report_name}' + '_' + str(num) + '.xlsx'
            try:
                with ExcelWriter(new_file_name,
                                 engine='openpyxl') as writer:
                    # проверяем тип объекта
                    # либо список
                    if isinstance(obj, list):
                        df = pd.DataFrame(obj).transpose()
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    # либо словарь
                    elif isinstance(obj, dict):
                        df = pd.DataFrame.from_dict(obj, orient='index')
                        df = df.transpose()
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            except IndexError:
                obj.to_excel(fr'C:\generation_results\{report_name}.xlsx',
                             sheet_name=sheet_name,
                             index=False)
    except Exception as error:
        return error
    return 'Запись в файл выполнена'


# @timemometr
def head_of_table(path, sheet_name: str = 'Лист1'):
    """Функция определяет расположение числовой шапки и возвращает координаты её начала,
    есть возможность вернуть значения этой шапки.

    @param path: путь к файлам
    @param sheet_name: наименование листа
    @return: шапку проверяемых файлов
    """
    head_row_col_index = []
    miss_files = []
    head_of_file = []
    none_list = []
    num_of_col = 1
    index_of_head_row = 0
    df = pd.ExcelFile(path)
    file_name = Path(df).name
    wb = op.load_workbook(path)
    # проверяем лист на его корректность
    try:
        ws = wb[sheet_name]
    # если указанного листа не существует, переходим на активный лист
    except KeyError:
        ws = wb.active
    # проверяем первый столбец по принципу, проверки первых 20 ячеек на наличие значений
    # если значения отсутствуют, значит первый столбец в файле пустой и производится сдвиг
    # на следующий столбец, до тех пор пока не будет найдене индекс столбца с данными
    while True:
        for col_empty in ws.iter_cols(min_col=num_of_col, max_col=num_of_col, max_row=20):
            for cell in col_empty:
                none_list.append(cell.value)
        if list(set(none_list))[0] is None:
            num_of_col += 1
            continue
        break
    # после полученя идекса столбца перехоим к определению начала шапки
    for col in ws.iter_cols(min_col=num_of_col, max_col=num_of_col):
        for cell in col:
            # если ячейка пустая пропускаем её
            if cell.value is None:
                pass
            else:
                try:
                    # если не пустая, пробуем привести тип данных ячейки к числовому
                    int_cell = int(cell.value)
                    # если все получилось, проверяем ячейку на наличие 1 в ней
                    if int_cell == 1:
                        # если 1 найдена, прибавляем 1 к счетчику идекса колонки,
                        # что бы перейти к следующей колонке
                        num_of_col += 1
                        # в следующей колонке повторяем операции, но теперь ищем ячейку со значеним 2
                        for col_second in ws.iter_cols(min_col=num_of_col, max_col=num_of_col):
                            for cell_second in col_second:
                                if cell_second.value is None:
                                    pass
                                else:
                                    try:
                                        # двойная проверка, на наличие в столбце значений 02
                                        int_cell = int(cell_second.value)
                                        str_cell = str(cell_second.value)
                                        # если попадется цифра 02, она будет отсеена
                                        if len(str_cell) == 1:
                                            if int_cell == 2:
                                                # пройдя все провекри и найди 1 и 2, мы можем узнать координаты шапки
                                                star_index_of_row = cell.row
                                                star_index_of_col = cell.column
                                                head_row_col_index.append(star_index_of_row)
                                                head_row_col_index.append(star_index_of_col)
                                                # извлекаем индекс строки содержащей шапку,
                                                # передавая его в пустой счетчик
                                                index_of_head_row += star_index_of_row
                                            else:
                                                break
                                        else:
                                            pass
                                    except ValueError:
                                        pass
                # если не получилось привести тип данных ячейки к числовому,
                # получаем ошбику и пропускаем эту ячейку
                except ValueError:
                    pass
        # получаем всю строку с шапкой по полученному индексу
        for row in ws.iter_rows(min_row=index_of_head_row, max_row=index_of_head_row):
            try:
                for cell in row:
                    if cell.value is None:
                        pass
                    else:
                        # и записываем эту строчку в список
                        head_of_file.append(cell.value)
            except ValueError:
                miss_files.append(file_name)
                return miss_files
        result = {tuple(head_row_col_index): head_of_file}
        return result


# @timemometr
# def n():
#     l = []
#     for v in file_path:
#         l.append(head_of_table(v))
#     return l
#

# print(n())
